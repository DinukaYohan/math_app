# db.py — SQLite helpers + schema used by auth.py and app.py
import json
import os, sqlite3
from typing import Iterable, Optional
from flask import g
import uuid
from openpyxl import load_workbook

# DB file lives next to this module
DB_PATH = os.path.join(os.path.dirname(__file__), "app.db")

def _connect():
     # open SQLite with a small lock timeout + allow threaded use (Flask)
    con = sqlite3.connect(DB_PATH, timeout=5.0, check_same_thread=False)
    con.row_factory = sqlite3.Row
     # WAL = better concurrency; keep durability reasonable
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=NORMAL;")
    # enforce foreign keys at the DB level
    con.execute("PRAGMA foreign_keys=ON;")
    return con

def get_db():
     # one connection per request, cached on Flask 'g'
    if "db" not in g:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True) # ensure folder
        g.db = _connect()
    return g.db

def close_db(_exc=None):
     # close at request teardown
    db = g.pop("db", None)
    if db: db.close()

def init_db():
     # create tables/indexes if missing
     # created a table learning_objectives for the data collected from the excel sheet
    db = _connect()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      email TEXT UNIQUE NOT NULL,
      username TEXT UNIQUE NOT NULL,
      password_hash TEXT NOT NULL,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS qa_pairs(
      qaid TEXT PRIMARY KEY,
      user_id INTEGER NOT NULL,
      question TEXT NOT NULL,
      answer   TEXT NOT NULL,
      model    TEXT,
      meta_json TEXT,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_qa_user_created
      ON qa_pairs(user_id, created_at DESC);
    CREATE TABLE IF NOT EXISTS learning_objectives (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      country TEXT NOT NULL,
      grade   TEXT NOT NULL,
      language TEXT NOT NULL,
      topic    TEXT NOT NULL,
      objective TEXT NOT NULL,
      UNIQUE(country, grade, language, topic, objective)
    );
    CREATE INDEX IF NOT EXISTS idx_lo_cgl 
      ON learning_objectives(country, grade, language);
    CREATE INDEX IF NOT EXISTS idx_lo_cglt 
      ON learning_objectives(country, grade, language, topic);
    """)
    db.commit()
    db.close()

# --- functions expected by auth.py ---
def create_user(email, username, password_hash):
    # insert user; returns new integer id
    db = get_db()
    cur = db.execute(
        "INSERT INTO users(email, username, password_hash) VALUES(?,?,?)",
        (email, username, password_hash)
    )
    db.commit()
    return cur.lastrowid

def get_user_by_email(email):
        # fetch single user by email
    return get_db().execute(
        "SELECT id, email, username, password_hash, created_at FROM users WHERE email=?",
        (email,)
    ).fetchone()

def get_user_by_username(username):
    # fetch single user by username
    return get_db().execute(
        "SELECT id, email, username, password_hash, created_at FROM users WHERE username=?",
        (username,)
    ).fetchone()

def get_user_by_id(uid):
        # fetch single user by id
    return get_db().execute(
        "SELECT id, email, username, password_hash, created_at FROM users WHERE id=?",
        (uid,)
    ).fetchone()

#Helper functions used by /generate

def save_qa(user_id: int, question: str, answer: str, model: str, meta: dict | None = None) -> str:
    qaid = str(uuid.uuid4())
    db = get_db()
    db.execute(
        "INSERT INTO qa_pairs (qaid, user_id, question, answer, model, meta_json) VALUES (?,?,?,?,?,?)",
        (qaid, user_id, question, answer, model, json.dumps(meta or {})),
    )
    db.commit()
    return qaid



def list_qa_for_user(user_id: int, limit: int = 20, offset: int = 0):
    return get_db().execute(
        """
        SELECT qaid, question, answer, model, meta_json, created_at
        FROM qa_pairs
        WHERE user_id=?
        ORDER BY datetime(created_at) DESC
        LIMIT ? OFFSET ?
        """,
        (user_id, limit, offset),
    ).fetchall()


def get_qa(qaid: str, user_id: int) -> Optional[sqlite3.Row]:
    #Fetch a single questions and answers item by its id, scoped to the owner.
    return (
        get_db()
        .execute(
            """
            SELECT qaid, user_id, question, answer, model, created_at
            FROM qa_pairs
            WHERE qaid=? AND user_id=?
            """,
            (qaid, user_id),
        )
        .fetchone()
    )

def delete_qa(qaid: str, user_id: int) -> int:
    #Delete a questions and answers item by id, scoped to the owner. Returns number of rows deleted (0 or 1).
    db = get_db()
    cur = db.execute(
        "DELETE FROM qa_pairs WHERE qaid=? AND user_id=?",
        (qaid, user_id),
    )
    db.commit()
    return cur.rowcount

# --------- Data from the excel file importer and utilities ----------

#cleans up text values from excel so they’re safe and consistent.
def _canon(s: str) -> str:
    """Trim/normalize a cell value into a safe string."""
    return (s or "").strip()

#Reads the excel file, cleans each row and saves all learning objectives to the database 
def import_learning_objectives_xlsx(xlsx_path: str) -> int:
    """
    Load/refresh learning objectives from an Excel file into SQLite.
    This is idempotent: duplicates are ignored via UNIQUE constraint.
    Returns the number of rows processed (attempted).
    Expected columns (case-insensitive):
      Country | Grade | Language | Topic | Learning Objective
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"LO spreadsheet not found: {xlsx_path}")

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [(_canon(c.value)).lower() for c in header_cells]

    required = ["country", "grade", "language", "topic", "learning objective"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError(
            f"Spreadsheet missing required columns: {', '.join(missing)}. "
            "Expected columns: Country, Grade, Language, Topic, Learning Objective"
        )

    i_country = headers.index("country")
    i_grade   = headers.index("grade")
    i_lang    = headers.index("language")
    i_topic   = headers.index("topic")
    i_obj     = headers.index("learning objective")

    db = get_db()
    n = 0
    for row in ws.iter_rows(min_row=2):
        country = _canon(row[i_country].value)
        grade   = _canon(str(row[i_grade].value))
        lang    = _canon(row[i_lang].value)
        topic   = _canon(row[i_topic].value)
        obj     = _canon(row[i_obj].value)

        # Skip incomplete rows
        if not (country and grade and lang and topic and obj):
            continue

        # Insert or ignore duplicates
        db.execute(
            """INSERT OR IGNORE INTO learning_objectives
               (country, grade, language, topic, objective)
               VALUES (?,?,?,?,?)""",
            (country, grade, lang, topic, obj),
        )
        n += 1

    db.commit()
    return n

#Asks the datavase for all different countries that exists
def list_distinct_countries():
    return [r[0] for r in get_db().execute(
        "SELECT DISTINCT country FROM learning_objectives ORDER BY country COLLATE NOCASE"
    ).fetchall()]

# Asks the database for all languages filtered by a specific country and grade
def list_distinct_languages(country=None, grade=None):
    q  = "SELECT DISTINCT language FROM learning_objectives WHERE 1=1"
    ps = []
    if country:
        q += " AND country=?"; ps.append(country)
    if grade:
        q += " AND grade=?";   ps.append(grade)
    q += " ORDER BY language COLLATE NOCASE"
    return [r[0] for r in get_db().execute(q, ps).fetchall()]

# Asks the database for all grades filtered by a specfic language and country
def list_distinct_grades(country=None, language=None):
    q  = "SELECT DISTINCT grade FROM learning_objectives WHERE 1=1"
    ps = []
    if country:
        q += " AND country=?";  ps.append(country)
    if language:
        q += " AND language=?"; ps.append(language)
    q += " ORDER BY CAST(grade AS INTEGER)"
    return [r[0] for r in get_db().execute(q, ps).fetchall()]

# Asks the database for all topics based on the country, grade and language
def list_topics(country, grade, language):
    return [r[0] for r in get_db().execute(
        """SELECT DISTINCT topic FROM learning_objectives
           WHERE country=? AND grade=? AND language=?
           ORDER BY topic COLLATE NOCASE""",
        (country, grade, language)
    ).fetchall()]

# Looks for all learning objectives in the database that match the selection of  country, grade, language and topic
def list_objectives(country, grade, language, topic):
    return [r[0] for r in get_db().execute(
        """SELECT objective FROM learning_objectives
           WHERE country=? AND grade=? AND language=? AND topic=?
           ORDER BY objective COLLATE NOCASE""",
        (country, grade, language, topic)
    ).fetchall()]

# Checks if the specific combination of the selected options actually exists
def combo_is_valid(country, grade, language, topic, objective=None) -> bool:
    if objective:
        row = get_db().execute(
            """SELECT 1 FROM learning_objectives
               WHERE country=? AND grade=? AND language=? AND topic=? AND objective=?
               LIMIT 1""",
            (country, grade, language, topic, objective)
        ).fetchone()
    else:
        row = get_db().execute(
            """SELECT 1 FROM learning_objectives
               WHERE country=? AND grade=? AND language=? AND topic=?
               LIMIT 1""",
            (country, grade, language, topic)
        ).fetchone()
    return bool(row)
